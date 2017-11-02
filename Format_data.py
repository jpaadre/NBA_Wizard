import openpyxl
from openpyxl import load_workbook
import os
from datetime import date
import pandas as pd
from pandas import ExcelWriter
os.chdir('C:\\Users\\Peter Haley\\Desktop\\Projects\\Data_Science\\Python\\NBA\\Excel')

# year = '2016'


teamList = ['ATL','BOS','BKN','CHA','CHI','CLE','DAL','DEN','DET','GSW','HOU',
         'IND','LAC','LAL','MEM','MIA','MIL','MIN','NOP','NYK','OKC','ORL',
         'PHI','PHX','POR','SAC','SAS','TOR','UTA','WAS']


def getDataSet(dataset):
    df = pd.read_excel(dataset)
    print('Dataset Loaded')
    return df

def getDataSetcsv(dataset):
    df = pd.read_csv(dataset)
    print('Dataset Loaded')
    return df

def getDateCutoff(year):
    # year = int(year)
    if year == '2017':
        return '2017-10-17'
    if year == '2016':
        return '2016-10-25'
    if year == '2015':
        return '2015-10-27'
    else:
        return('Invalid Year')

def SplitTeams(df,year):
    writer = ExcelWriter('AllStats_' + year + '_Split.xlsx')
    print('Splitting Data')
    cutoff = getDateCutoff(year)
    print(cutoff)

    df['GAME_DATE'] = pd.to_datetime(df['GAME_DATE_EST'])
    df = df.loc[df['GAME_DATE'] >= cutoff]
    for i in teamList:
        df1 = df.loc[df['TEAM_ABBREVIATION'] == i]
        #Format Date and calc rest
        df1['GAME_DATE'] = pd.to_datetime(df1['GAME_DATE_EST'])
        df1['DaysRest'] = df1['GAME_DATE'] - df1['GAME_DATE'].shift(1)
        df1['HomeTeam'] = df1['GAMECODE'].str[12:]
        df1['AwayTeam'] = df1['GAMECODE'].str[9:12]
        df1['HomeIndex'] = df1.apply(getHomeIndex,axis=1)
        df1['AvgPace']= df1['PACE'].expanding().mean()
        df1['AvgORTG'] = df1['OFF_RATING'].expanding().mean()
        df1['AvgDRTG'] = df1['DEF_RATING'].expanding().mean()
        df1['AvgORTG_L5'] = df1['OFF_RATING'].rolling(window=5).mean()
        df1['AvgDRTG_L5'] = df1['DEF_RATING'].rolling(window=5).mean()
        df1['AvgPace']= df1['AvgPace'].shift(1)
        df1['AvgORTG'] = df1['AvgORTG'].shift(1)
        df1['AvgDRTG'] = df1['AvgDRTG'].shift(1)
        df1['AvgORTG_L5'] = df1['AvgORTG_L5'].shift(1)
        df1['AvgDRTG_L5'] = df1['AvgDRTG_L5'].shift(1)
        df1['HomeORTG'] = df1.apply(getHomeORTG,axis=1)
        df1['AwayORTG'] = df1.apply(getAwayORTG,axis=1)
        df1['HomeDRTG'] = df1.apply(getHomeDRTG,axis=1)
        df1['AwayDRTG'] = df1.apply(getAwayDRTG,axis=1)
        df1['HomeORTG'] = df1['HomeORTG'].expanding().mean()
        df1['AwayORTG'] = df1['AwayORTG'].expanding().mean()
        df1['HomeDRTG'] = df1['HomeDRTG'].expanding().mean()
        df1['AwayDRTG'] = df1['AwayDRTG'].expanding().mean()
        df1['HomeORTG'] = df1['HomeORTG'].shift(1)
        df1['AwayORTG'] = df1['AwayORTG'].shift(1)
        df1['HomeDRTG'] = df1['HomeDRTG'].shift(1)
        df1['AwayDRTG'] = df1['AwayDRTG'].shift(1)
        df1['Location_Avg_ORTG'] = df1.apply(getLocationORTG,axis=1)
        df1['Location_Avg_DRTG'] = df1.apply(getLocationDRTG,axis=1)

        df1.to_excel(writer,i)
    writer.save()
    print('Data Split')


def getFirstSplit(year):
    fileName = ('AllStats_'+ year +'_Split.xlsx')
    wb = load_workbook(fileName)
    df = pd.read_excel(fileName, sheetname='ATL')
    tabs = wb.get_sheet_names()
    print("Extracting Team Data")
    for j in tabs:
        if j != 'ATL':
            df4 = pd.read_excel(fileName, sheetname=j)
            frames = [df,df4]
            df= pd.concat(frames)
    dfH = df[(df['HomeIndex'] == 0)]
    dfA = df[(df['HomeIndex'] == 1)]
    df1 = pd.merge(dfH, dfA, on='GAMECODE',how='outer')
    df2 = pd.merge(dfA, dfH, on='GAMECODE',how='outer')
    dfList = [df1,df2]
    df3= pd.concat(dfList)
    # print(df3.head())
    df3 = df3.sort_values(by=['GAMECODE','HomeIndex_x'],ascending=[True,True])

    print("Complete")
    return df3



def getHomeIndex(df):
    if df['HomeTeam'] == df['TEAM_ABBREVIATION']:
        return (0)
    else:
        return (1)

def getHomeORTG(df):
    if df['HomeIndex'] == 0:
        return df['OFF_RATING']


def getAwayORTG(df):
    if df['HomeIndex'] == 1:
        return df['OFF_RATING']


def getHomeDRTG(df):
    if df['HomeIndex'] == 0:
        return df['DEF_RATING']


def getAwayDRTG(df):
    if df['HomeIndex'] == 1:
        return df['DEF_RATING']

def getLocationORTG(df):
    if df['HomeIndex'] == 0:
        return df['HomeORTG']
    else:
        return df['AwayORTG']

def getLocationDRTG(df):
    if df['HomeIndex'] == 0:
        return df['HomeDRTG']
    else:
        return df['AwayDRTG']

def SplitJointTeams(df,year):
    writer = ExcelWriter('Split_Teams_2nd_Time_'+ year +'.xlsx')
    print('Splitting Data')
    for i in teamList:
        df1 = df.loc[df['TEAM_ABBREVIATION_x'] == i]
        #Format Date and calc rest
        df1['AvgPace_OPP']= df1['AvgPace_y'].expanding().mean()
        df1['AvgORTG_OPP'] = df1['AvgORTG_y'].expanding().mean()
        df1['AvgDRTG_OPP'] = df1['AvgDRTG_y'].expanding().mean()
        df1['AvgORTG_L5_OPP'] = df1['AvgORTG_y'].rolling(window=5).mean()
        df1['AvgDRTG_L5_OPP'] = df1['AvgDRTG_y'].rolling(window=5).mean()
        df1['Opp_ORTG_vs_Avg'] = df1['LeagueAvgORTG'] / df1['AvgORTG_OPP']
        df1['Opp_DRTG_vs_Avg'] = df1['LeagueAvgDRTG']/ df1['AvgDRTG_OPP']
        df1['Opp_Pace_vs_Avg'] = df1['AvgPace_OPP'] / df1['LeagueAvgPace']
        df1['Opp_ORTG_vs_Avg_L5'] = df1['LeagueAvgORTG_L5'] / df1['AvgORTG_L5_OPP']
        df1['Opp_DRTG_vs_Avg_L5'] = df1['LeagueAvgDRTG_L5']/ df1['AvgDRTG_L5_OPP']

        df1['std_AvgORTG'] = df1['AvgORTG_x'] * df1['Opp_DRTG_vs_Avg']
        df1['std_AvgDRTG'] = df1['AvgDRTG_x'] / df1['Opp_ORTG_vs_Avg']
        df1['std_AvgORTG_L5'] = df1['AvgORTG_L5_x'] * df1['Opp_DRTG_vs_Avg_L5']
        df1['std_AvgDRTG_L5'] = df1['AvgDRTG_L5_x'] / df1['Opp_ORTG_vs_Avg_L5']


        df1.to_excel(writer,i)

    writer.save()
    return(df1)
    print('Data Split')

def getSecondSplit(year):
    fileName = ('Split_Teams_2nd_Time_'+ year +'.xlsx')
    wb = load_workbook(fileName)
    df = pd.read_excel(fileName, sheetname='ATL')
    tabs = wb.get_sheet_names()
    print("Extracting Team Data")
    for j in tabs:
        if j != 'ATL':
            df1 = pd.read_excel(fileName, sheetname=j)
            frames = [df,df1]
            df= pd.concat(frames)

    df1 =df[['GAMECODE','HomeIndex_x','std_AvgORTG', 'std_AvgDRTG','std_AvgORTG_L5','std_AvgDRTG_L5']]
    df1 = df1.sort_values(by=['GAMECODE','HomeIndex_x'],ascending=[True,True])
    # print(df1.head())
    df2 = pd.merge(df, df1, on=['GAMECODE','HomeIndex_x'])
    # print(df2.head())
    df2 = df2.sort_values(by=['GAMECODE','HomeIndex_x'],ascending=[True,True])
    # print(df2.head())
    df2 = df2[['GAMECODE','TEAM_ABBREVIATION_x','HomeIndex_x','DaysRest_x','AvgPace_x','AvgORTG_x','AvgDRTG_x','AvgORTG_L5_x','AvgDRTG_L5_x','std_AvgORTG_x', 'std_AvgDRTG_x','std_AvgORTG_L5_x',
    'std_AvgDRTG_L5_x','TEAM_ABBREVIATION_y','DaysRest_y','AvgPace_y','AvgORTG_y','AvgDRTG_y','AvgORTG_L5_y','AvgDRTG_L5_y','HomeORTG_x','HomeDRTG_x','AwayORTG_x','AwayDRTG_x','Location_Avg_ORTG_x','Location_Avg_DRTG_x',
    'HomeORTG_y','HomeDRTG_y','AwayORTG_y','AwayDRTG_y','Location_Avg_ORTG_y','Location_Avg_DRTG_y','std_AvgORTG_y', 'std_AvgDRTG_y','std_AvgORTG_L5_y','std_AvgDRTG_L5_y','OFF_RATING_x']]
    #
    # writer = ExcelWriter("testMerge.xlsx")
    # df1.to_excel(writer,'Master')
    # writer.save()
    writer1 = ExcelWriter('DataForModel_'+ year +'.xlsx')
    df2.to_excel(writer1,'Master')
    writer1.save()
    return df2

def getLeagueAvg(df,year):
    df['GAMELINK'] = df['GAMECODE'].str[:8]
    df1 = df[['OFF_RATING_x','GAMELINK']]
    df2 = df[['DEF_RATING_x','GAMELINK']]
    df3 = df[['PACE_x','GAMELINK']]
    df1 = df1.groupby(['GAMELINK'],as_index=False)['OFF_RATING_x'].mean()
    df2 = df2.groupby(['GAMELINK'],as_index=False)['DEF_RATING_x'].mean()
    df3 = df3.groupby(['GAMELINK'],as_index=False)['PACE_x'].mean()

    df4 = pd.merge(df1, df2, on='GAMELINK',how='outer')
    df5 = pd.merge(df4, df3, on='GAMELINK',how='outer')
    df5['LeagueAvgORTG'] = df5['OFF_RATING_x'].expanding().mean()
    df5['LeagueAvgDRTG'] = df5['DEF_RATING_x'].expanding().mean()
    df5['LeagueAvgPace'] = df5['PACE_x'].expanding().mean()
    df5['LeagueAvgORTG'] = df5['LeagueAvgORTG'].shift(1)
    df5['LeagueAvgDRTG'] = df5['LeagueAvgDRTG'].shift(1)
    df5['LeagueAvgPace'] = df5['LeagueAvgPace'].shift(1)
    #
    df5 = df5[['GAMELINK','LeagueAvgORTG','LeagueAvgDRTG','LeagueAvgPace']]
    df5['LeagueAvgORTG_L5'] = df5['LeagueAvgORTG'].rolling(window=5).mean()
    df5['LeagueAvgDRTG_L5'] = df5['LeagueAvgDRTG'].rolling(window=5).mean()
    #
    df6 = pd.merge(df, df5, on='GAMELINK',how='outer')

    # print(df6.head())
    writer1 = ExcelWriter('LeagueAvg_'+ year + '.xlsx')
    df5.to_excel(writer1,'Master')
    writer1.save()
    return df6


def CalcStats(year):
    # ---------------- Split data out by team and calculate stats ------------------------
    dataset = getDataSet('AllStats_'+ year + '.xlsx')
    SplitTeams(dataset,year)
    # ---------------- Consume team data from tabs to make one dataset and combine both teams onto one line------------------------
    teamdata1 = getFirstSplit(year)
    teamdata2 = getLeagueAvg(teamdata1,year)
    # #---------------- Split data out again to calculate Opponent stats ------------------------
    teamdata3 = SplitJointTeams(teamdata2,year)
    # #---------------- Cosolidate split team data, remove first 6 rows without Last 5 calcs, Filter only needed columns------------------------
    teamdata4 = getSecondSplit(year)
    print('Stat Calculation Complete')
