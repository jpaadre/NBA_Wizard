import openpyxl
from openpyxl import load_workbook
import os
from datetime import date
import getAdjDRTG
os.chdir('c:\\PythonScripts\\Excel')

wb = load_workbook('Team_Stats_Raw.xlsx')

teamIDs ={
 1610612737: 'ATL',
 1610612738: 'BOS',
 1610612751: 'BKN',
 1610612766: 'CHA',
 1610612741: 'CHI',
 1610612739: 'CLE',
 1610612742: 'DAL',
 1610612743: 'DEN',
 1610612765: 'DET',
 1610612744: 'GSW',
 1610612745: 'HOU',
 1610612754: 'IND',
 1610612746: 'LAC',
 1610612747: 'LAL',
 1610612763: 'MEM',
 1610612748: 'MIA',
 1610612749: 'MIL',
 1610612750: 'MIN',
 1610612740: 'NOP',
 1610612752: 'NYK',
 1610612760: 'OKC',
 1610612753: 'ORL',
 1610612755: 'PHI',
 1610612756: 'PHX',
 1610612757: 'POR',
 1610612758: 'SAC',
 1610612759: 'SAS',
 1610612761: 'TOR',
 1610612762: 'UTA',
 1610612764: 'WAS'
 }


def getOpponentAvg(todaysGames):
    allRTGs = {}
    for i in todaysGames:
        homeTeam = i[3:7]
        awayTeam = i[0:3]
        games= []
        games.append(awayTeam)
        games.append(homeTeam)
        for t in games:
            row_cell =2
            tab = wb.get_sheet_by_name(t)
            row_max = tab.max_row
            hometeam_oppRTG = []
            awayteam_oppRTG = []
            for j in range(row_max -1):
                tab = wb.get_sheet_by_name(t)
                row_max = tab.max_row
                location = tab.cell(row = row_cell, column = 8).value
                if location == 'Home':
                    oppID = tab.cell(row = row_cell, column = 6).value
                    opp = teamIDs[oppID]
                    tab = wb.get_sheet_by_name('DRTG')
                    summary_max_row = tab.max_row
                    summary_row_cell = 2
                    summary_column_cell = 1
                    for m in range(summary_max_row -1):
                        if opp == tab.cell(row = summary_row_cell, column = summary_column_cell).value:
                            rtg_row = summary_row_cell
                        summary_row_cell +=1
                    rtg = tab.cell(row = rtg_row, column = 18).value
                    hometeam_oppRTG.append(rtg)
                else:
                    oppID = tab.cell(row = row_cell, column = 5).value
                    opp = teamIDs[oppID]
                    tab = wb.get_sheet_by_name('DRTG')
                    summary_max_row = tab.max_row
                    summary_row_cell = 2
                    summary_column_cell = 1
                    for m in range(summary_max_row -1):
                        if opp == tab.cell(row = summary_row_cell, column = summary_column_cell).value:
                            rtg_row = summary_row_cell
                        summary_row_cell +=1
                    rtg = tab.cell(row = rtg_row, column = 4).value
                    hometeam_oppRTG.append(rtg)
                row_cell +=1
            count = 0
            total = 0
            for a in hometeam_oppRTG:
               count +=1
               total = total + a
            oppDRTG = total/count
            allRTGs[t] = oppDRTG
    #for h in allRTGs:
        #print (h, allRTGs[h])
    return allRTGs
           
           


def getAdjORTG(todaysGames,adjRatio):
    adjORTGS = {}
    for i in todaysGames:
        homeTeam = i[3:7]
        awayTeam = i[0:3]
        teams = []
        teams.append(awayTeam)
        teams.append(homeTeam)
        count = 0
        for t in teams:
            tab = wb.get_sheet_by_name('ORTG')
            summary_max_row = tab.max_row
            summary_row_cell = 2
            summary_column_cell = 1
            for m in range(summary_max_row -1):
                if t == tab.cell(row = summary_row_cell, column = summary_column_cell).value:
                    if count == 0:
                        ORTG = tab.cell(row = summary_row_cell, column = 18).value
                        rtg = ORTG * adjRatio[t]
                        adjORTGS[t] = rtg
                    else:
                        ORTG = tab.cell(row = summary_row_cell, column = 4).value
                        rtg = ORTG * adjRatio[t]
                        adjORTGS[t] = rtg
                summary_row_cell += 1
            count += 1
    #for h in adjORTGS:
        #print (h, adjORTGS[h])
    return adjORTGS



def getdefRatios(todaysGames):
    DRTGs = {}
    for i in todaysGames:
        homeTeam = i[3:7]
        awayTeam = i[0:3]
        teams = []
        teams.append(awayTeam)
        teams.append(homeTeam)
        count = 0
        tab = wb.get_sheet_by_name('DRTG')
        summary_max_row = tab.max_row
        for x in teams:
            summary_row_cell = 2
            summary_column_cell = 1
            for m in range(summary_max_row -1):
                if x == tab.cell(row = summary_row_cell, column = summary_column_cell).value:
                    rtg_row = summary_row_cell
                    if count == 0:
                        teamDRTG = tab.cell(row = rtg_row, column = 18).value
                    else:
                        teamDRTG = tab.cell(row = rtg_row, column = 4).value
                    DRTGs[x] = teamDRTG
                summary_row_cell +=1
            count +=1
    #for h in DRTGs:
        #print (h, DRTGs[h])
    return DRTGs

#rtg = getOpponentAvg(todaysGames)        
#DRTGs = getdefRatios(todaysGames)

def getAdjRatio(todaysGames,DRTGs, avgfaced):
    adjORTG = {}
    for i in todaysGames:
        homeTeam = i[3:7]
        awayTeam = i[0:3]
        teams = []
        teams.append(awayTeam)
        teams.append(homeTeam)
        count = 0
        for j in teams:
            if count == 0:
                ratio = DRTGs[homeTeam]/avgfaced[j]
                adjORTG[j] = ratio
            else:
                ratio = DRTGs[awayTeam]/avgfaced[j]
                adjORTG[j] = ratio
            count +=1
    #for h in adjORTG:
        #print (h, adjORTG[h])
    return adjORTG


def ProjectScores(todaysGames, ORTGs):
    proj = {}
    for i in todaysGames:
        homeTeam = i[3:7]
        awayTeam = i[0:3]
        teams = []
        teams.append(awayTeam)
        teams.append(homeTeam)
        tab = wb.get_sheet_by_name('Pace')
        for x in teams:
            row_cell = 2
            row_max = tab.max_row
            for j in range(row_max -1):
                if homeTeam == tab.cell(row = row_cell, column = 1).value:
                    homePace = tab.cell(row = row_cell, column = 4).value
                if awayTeam == tab.cell(row = row_cell, column = 1).value:
                    awayPace = tab.cell(row = row_cell, column = 18).value
                row_cell +=1
            avgPace = (homePace + awayPace) / 2
            projection = (avgPace * ORTGs[x])/100
            proj[x] = projection
    for g in proj:
        print(g)
    for h in proj:
        print(proj[h])
    return proj
              
#adjRatio = getAdjRatio(todaysGames,DRTGs,rtg)
#getAdjORTG(todaysGames,adjRatio)

#Standardized vs non standard
#test1 = getAdjDRTG.getDRTG(todaysGames)
#test2 = getdefRatios(todaysGames)
def getORTG(todaysGames):
    rtg = getOpponentAvg(todaysGames)        
    DRTGs = getAdjDRTG.getDRTG(todaysGames)
    adjRatio = getAdjRatio(todaysGames,DRTGs,rtg)
    adjRTG = getAdjORTG(todaysGames,adjRatio)
    return adjRTG
