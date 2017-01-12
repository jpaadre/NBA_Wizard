import openpyxl
from openpyxl import load_workbook
import os
from datetime import date
os.chdir('c:\\PythonScripts\\Excel')

wb = load_workbook('Team_Stats_Raw.xlsx')


teamIDs = {'ATL':0,'BOS':0,'BKN':0,'CHA':0,'CHI':0,'CLE':0,'DAL':0,'DEN':0,'DET':0,
         'GSW':0,'HOU':0,
         'IND':0,'LAC':0,'LAL':0,'MEM':0,'MIA':0,'MIL':0,'MIN':0,'NOP':0,'NYK':0,
         'OKC':0,'ORL':0,
         'PHI':0,'PHX':0,'POR':0,'SAC':0,'SAS':0,'TOR':0,'UTA':0,'WAS':0}

teams = ['ATL','BOS','BKN','CHA','CHI','CLE','DAL','DEN','DET','GSW','HOU',
         'IND','LAC','LAL','MEM','MIA','MIL','MIN','NOP','NYK','OKC','ORL',
         'PHI','PHX','POR','SAC','SAS','TOR','UTA','WAS']

def getTeamIDs():
    for i in teams:
        tab = wb.get_sheet_by_name(i)
        teamIDs[i] = tab.cell(row = 2, column = 4).value

    for x in teamIDs:
        print(x, teamIDs[x])


def getLeagueAvg():
    summaryTabs = ['ORTG','DRTG','Pace']
    for j in summaryTabs:
        tab = wb.get_sheet_by_name(j)
        column_cell = 2
        row_max = tab.max_row
        column_max = tab.max_column
        for x in range(column_max-1):
            total = 0
            row_cell = 2
            count = 0
            for m in range(row_max-2):
                if tab.cell(row = row_cell, column = column_cell).value != None and tab.cell(row = row_cell, column = column_cell).value != 0:
                    total = total + tab.cell(row = row_cell, column = column_cell).value
                    count += 1
                row_cell +=1
            if count != 0:
                total = total /count
                tab.cell(row = row_max, column = column_cell).value = total
            column_cell +=1
    wb.save('Team_Stats_Raw.xlsx')

def getOpponentAvg(team):
    tab = wb.get_sheet_by_name(team)
    row_cell = 2
    column_cell = 2
    row_max = tab.max_row
    column_max = tab.max_column


        
    
    
    
