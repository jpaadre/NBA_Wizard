import openpyxl
from openpyxl import Workbook 

import os
os.chdir('c:\\PythonScripts\\Excel')

wb = Workbook()
dest_filename = 'Team_Stats_Raw.xlsx' 



Teams = ['Master','ORTG','ATL','BOS','BKN','CHA','CHI','CLE','DAL','DEN','DET','GSW','HOU',
         'IND','LAC','LAL','MEM','MIA','MIL','MIN','NOP','NYK','OKC','ORL',
         'PHI','PHO','POR','SAC','SAS','TOR','UTA','WAS']

columns = ['GameID','Date','GameCode','TeamID','HomeID','AwayID','Team',
           'Location','ORTG','DRTG','Pace','PIE','Rest']
headers = ['Total_ORTG','Total_ORTG_count','Home_ORTG','Home_ORTG_count',
           'Home_ORTG_r0','Home_ORTG_r0_count','Home_ORTG_r1',
           'Home_ORTG_r1_count','Home_ORTG_r2','Home_ORTG_r2_count',
           'Home_ORTG_r3','Home_ORTG_r3_count','Home_ORTG_r4',
           'Home_ORTG_r4_count','Home_ORTG_3N4','Home_ORTG_3N4_count',
           'Home_ORTG_4N5','Home_ORTG_4N5_count','Away_ORTG',
           'Away_ORTG_count','Away_ORTG_r0','Away_ORTG_r0_count',
           'Away_ORTG_r1','Away_ORTG_r1_count','Away_ORTG_r2',
           'Away_ORTG_r2_count','Away_ORTG_r3','Away_ORTG_r3_count',
           'Away_ORTG_r4','Away_ORTG_r4_count','Away_ORTG_3N4',
           'Away_ORTG_3N4_count','Away_ORTG_4N5','Away_ORTG_4N5_count']

def createNewBook():
    for i in Teams:
        wb.create_sheet(title=i)
        sheet = wb.get_sheet_by_name(i)
        row_cell = sheet.max_row
        column_cell = 1
        for j in columns:
            sheet.cell(row = row_cell, column = column_cell).value = j
            column_cell= column_cell + 1
    wb.save('Team_Stats_Raw1.xlsx')



sheet = wb.get_sheet_by_name('ORTG')
row_cell =2
column_cell = 1
for i in Teams:
    sheet.cell(row = row_cell, column = column_cell).value = i
    column_cell +=1

sheet = wb.get_sheet_by_name('ORTG')
row_cell =1
column_cell = 1
for k in headers:
    sheet.cell(row = row_cell, column = column_cell).value = k
    row_cell +=1
wb.save('Team_Stats_Raw.xlsx')
