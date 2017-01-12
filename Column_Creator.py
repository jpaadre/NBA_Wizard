import openpyxl
#from openpyxl import Workbook FOR CREATING FILE
from openpyxl import load_workbook
import os
os.chdir('c:\\PythonScripts\\Excel')

'''wb = Workbook()
dest_filename = 'Team_Stats_Raw1.xlsx' FOR CREATING WORKBOOK
''' 
wb = load_workbook(filename ='Team_Stats_Raw1.xlsx')

Teams = ['ATL','BOS','BKN','CHA','CHI','CLE','DAL','DEN','DET','GSW','HOU',
         'IND','LAC','LAL','MEM','MIA','MIL','MIN','NOP','NYK','OKC','ORL',
         'PHI','PHO','POR','SAC','SAS','TOR','UTA','WAS']

columns = ['GameID','Date','GameCode','TeamID','HomeID','AwayID','Team',
           'Location','ORTG','DRTG','Pace','PIE']


for i in Teams:
    sheet = wb.get_sheet_by_name(i)
    row_cell = sheet.max_row
    column_cell = 1
    for j in columns:
        sheet.cell(row = row_cell, column = column_cell).value = j
        column_cell= column_cell + 1



wb.save('Team_Stats_Raw1.xlsx')
