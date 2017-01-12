import openpyxl
from openpyxl import load_workbook
import os
from datetime import date
os.chdir('c:\\PythonScripts\\Excel')

wb = load_workbook('Team_Stats_Raw.xlsx')

def calcORTG():
    tabs = wb.get_sheet_names()
    for j in tabs:
        if j != 'Master' and j!= 'ORTG' and j != 'DRTG' and j!= 'Pace':
            tab = wb.get_sheet_by_name(j)
            row_cell = 2
            column_cell = 1
            column_count = tab.max_column
            row_count = tab.max_row
            Total_ORTG = 0
            Total_ORTG_count = 0
            Home_ORTG = 0
            Home_ORTG_count = 0
            Home_ORTG_r0 = 0
            Home_ORTG_r0_count = 0
            Home_ORTG_r1 = 0
            Home_ORTG_r1_count = 0
            Home_ORTG_r2 = 0
            Home_ORTG_r2_count = 0
            Home_ORTG_r3 = 0
            Home_ORTG_r3_count = 0
            Home_ORTG_3N4 = 0
            Home_ORTG_3N4_count = 0
            Home_ORTG_4N5 = 0
            Home_ORTG_4N5_count = 0
            Away_ORTG = 0
            Away_ORTG_count = 0
            Away_ORTG_r0 = 0
            Away_ORTG_r0_count = 0
            Away_ORTG_r1 = 0
            Away_ORTG_r1_count = 0
            Away_ORTG_r2 = 0
            Away_ORTG_r2_count = 0
            Away_ORTG_r3 = 0
            Away_ORTG_r3_count = 0
            Away_ORTG_3N4 = 0
            Away_ORTG_3N4_count = 0
            Away_ORTG_4N5 = 0
            Away_ORTG_4N5_count = 0
            statsList =[]
            ORTG_cell = 9
            Location_cell = 8
            Rest_cell = 13
            Rest_Detail_cell = 14
            for m in range(row_count-1):
                Total_ORTG = Total_ORTG + tab.cell(row = row_cell, column = ORTG_cell).value
                Total_ORTG_count = Total_ORTG_count + 1
                if tab.cell(row = row_cell, column = Location_cell).value == 'Home':
                    Home_ORTG = Home_ORTG + tab.cell(row = row_cell, column = ORTG_cell).value
                    Home_ORTG_count = Home_ORTG_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 0:
                        Home_ORTG_r0 = Home_ORTG_r0 + tab.cell(row = row_cell, column = ORTG_cell).value
                        Home_ORTG_r0_count = Home_ORTG_r0_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 1:
                        Home_ORTG_r1 = Home_ORTG_r1 + tab.cell(row = row_cell, column = ORTG_cell).value
                        Home_ORTG_r1_count = Home_ORTG_r1_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 2:
                        Home_ORTG_r2 = Home_ORTG_r2 + tab.cell(row = row_cell, column = ORTG_cell).value
                        Home_ORTG_r2_count = Home_ORTG_r2_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value >= 3:
                        Home_ORTG_r3 = Home_ORTG_r3 + tab.cell(row = row_cell, column = ORTG_cell).value
                        Home_ORTG_r3_count = Home_ORTG_r3_count + 1
                    if tab.cell(row = row_cell, column = Rest_Detail_cell).value == '3IN4':
                        Home_ORTG_3N4 = Home_ORTG_3N4 + tab.cell(row = row_cell, column = ORTG_cell).value
                        Home_ORTG_3N4_count = Home_ORTG_3N4_count + 1
                    if tab.cell(row = row_cell, column = Rest_Detail_cell).value == '4IN5':
                        Home_ORTG_4N5 = Home_ORTG_4N5 + tab.cell(row = row_cell, column = ORTG_cell).value
                        Home_ORTG_4N5_count = Home_ORTG_4N5_count + 1
                else:
                    Away_ORTG = Away_ORTG + tab.cell(row = row_cell, column = ORTG_cell).value
                    Away_ORTG_count = Away_ORTG_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 0:
                        Away_ORTG_r0 = Away_ORTG_r0 + tab.cell(row = row_cell, column = ORTG_cell).value
                        Away_ORTG_r0_count = Away_ORTG_r0_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 1:
                        Away_ORTG_r1 = Away_ORTG_r1 + tab.cell(row = row_cell, column = ORTG_cell).value
                        Away_ORTG_r1_count = Away_ORTG_r1_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 2:
                        Away_ORTG_r2 = Away_ORTG_r2 + tab.cell(row = row_cell, column = ORTG_cell).value
                        Away_ORTG_r2_count = Away_ORTG_r2_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value >= 3:
                        Away_ORTG_r3 = Away_ORTG_r3 + tab.cell(row = row_cell, column = ORTG_cell).value
                        Away_ORTG_r3_count = Away_ORTG_r3_count + 1
                    if tab.cell(row = row_cell, column = Rest_Detail_cell).value == '3IN4':
                        Away_ORTG_3N4 = Away_ORTG_3N4 + tab.cell(row = row_cell, column = ORTG_cell).value
                        Away_ORTG_3N4_count = Away_ORTG_3N4_count + 1
                    if tab.cell(row = row_cell, column = Rest_Detail_cell).value == '4IN5':
                        Away_ORTG_4N5 = Away_ORTG_4N5 + tab.cell(row = row_cell, column = ORTG_cell).value
                        Away_ORTG_4N5_count = Away_ORTG_4N5_count + 1
                row_cell = row_cell + 1
            if Total_ORTG_count >0:
                Total_ORTG = Total_ORTG/Total_ORTG_count
            if Home_ORTG_count >0:
                Home_ORTG = Home_ORTG/Home_ORTG_count
            if Home_ORTG_r0_count >0:
                Home_ORTG_r0 = Home_ORTG_r0/Home_ORTG_r0_count
            if Home_ORTG_r1_count >0:
                Home_ORTG_r1 = Home_ORTG_r1/Home_ORTG_r1_count
            if Home_ORTG_r2_count >0:
                Home_ORTG_r2 = Home_ORTG_r2/Home_ORTG_r2_count
            if Home_ORTG_r3_count >0:
                Home_ORTG_r3 = Home_ORTG_r3/Home_ORTG_r3_count
            if Home_ORTG_3N4_count >0:
                Home_ORTG_3N4 = Home_ORTG_3N4/Home_ORTG_3N4_count
            if Home_ORTG_4N5_count >0:
                Home_ORTG_4N5 = Home_ORTG_4N5/Home_ORTG_4N5_count
            if Away_ORTG_count >0:
                Away_ORTG = Away_ORTG/Away_ORTG_count
            if Away_ORTG_r0_count >0:
                Away_ORTG_r0 = Away_ORTG_r0/Away_ORTG_r0_count
            if Away_ORTG_r1_count >0:
                Away_ORTG_r1 = Away_ORTG_r1/Away_ORTG_r1_count
            if Away_ORTG_r2_count >0:
                Away_ORTG_r2 = Away_ORTG_r2/Away_ORTG_r2_count
            if Away_ORTG_r3_count >0:
                Away_ORTG_r3 = Away_ORTG_r3/Away_ORTG_r3_count
            if Away_ORTG_3N4_count >0:
                Away_ORTG_3N4 = Away_ORTG_3N4/Away_ORTG_3N4_count
            if Away_ORTG_4N5_count >0:
                Away_ORTG_4N5 = Away_ORTG_4N5/Away_ORTG_4N5_count
            statsList.append(Total_ORTG)
            statsList.append(Total_ORTG_count)
            statsList.append(Home_ORTG)
            statsList.append(Home_ORTG_count)
            statsList.append(Home_ORTG_r0)
            statsList.append(Home_ORTG_r0_count)
            statsList.append(Home_ORTG_r1)
            statsList.append(Home_ORTG_r1_count)
            statsList.append(Home_ORTG_r2)
            statsList.append(Home_ORTG_r2_count)
            statsList.append(Home_ORTG_r3)
            statsList.append(Home_ORTG_r3_count)
            statsList.append(Home_ORTG_3N4)
            statsList.append(Home_ORTG_3N4_count)
            statsList.append(Home_ORTG_4N5)
            statsList.append(Home_ORTG_4N5_count)
            statsList.append(Away_ORTG)
            statsList.append(Away_ORTG_count)
            statsList.append(Away_ORTG_r0)
            statsList.append(Away_ORTG_r0_count)
            statsList.append(Away_ORTG_r1)
            statsList.append(Away_ORTG_r1_count)
            statsList.append(Away_ORTG_r2)
            statsList.append(Away_ORTG_r2_count)
            statsList.append(Away_ORTG_r3)
            statsList.append(Away_ORTG_r3_count)
            statsList.append(Away_ORTG_3N4)
            statsList.append(Away_ORTG_3N4_count)
            statsList.append(Away_ORTG_4N5)
            statsList.append(Away_ORTG_4N5_count)
            team = j
            tab = wb.get_sheet_by_name('ORTG')
            summary_rows = tab.max_row + 1
            summary_row_cell = 2
            summary_column_cell = 1
            summary_print_column = 2
            print (team)
            for z in range(summary_rows):
                if team == tab.cell(row = summary_row_cell, column = summary_column_cell).value:
                    summary_print_cell = summary_row_cell
                summary_row_cell +=1
            for x in range(len(statsList)):
                tab.cell(row = summary_print_cell, column = summary_print_column).value = statsList[x]
                summary_print_column = summary_print_column + 1
    wb.save('Team_Stats_Raw.xlsx')


def calcDRTG():
    tabs = wb.get_sheet_names()
    for j in tabs:
        if j != 'Master' and j!= 'ORTG' and j != 'DRTG' and j!= 'Pace':
            tab = wb.get_sheet_by_name(j)
            row_cell = 2
            column_cell = 1
            column_count = tab.max_column
            row_count = tab.max_row
            Total_DRTG = 0
            Total_DRTG_count = 0
            Home_DRTG = 0
            Home_DRTG_count = 0
            Home_DRTG_r0 = 0
            Home_DRTG_r0_count = 0
            Home_DRTG_r1 = 0
            Home_DRTG_r1_count = 0
            Home_DRTG_r2 = 0
            Home_DRTG_r2_count = 0
            Home_DRTG_r3 = 0
            Home_DRTG_r3_count = 0
            Home_DRTG_3N4 = 0
            Home_DRTG_3N4_count = 0
            Home_DRTG_4N5 = 0
            Home_DRTG_4N5_count = 0
            Away_DRTG = 0
            Away_DRTG_count = 0
            Away_DRTG_r0 = 0
            Away_DRTG_r0_count = 0
            Away_DRTG_r1 = 0
            Away_DRTG_r1_count = 0
            Away_DRTG_r2 = 0
            Away_DRTG_r2_count = 0
            Away_DRTG_r3 = 0
            Away_DRTG_r3_count = 0
            Away_DRTG_3N4 = 0
            Away_DRTG_3N4_count = 0
            Away_DRTG_4N5 = 0
            Away_DRTG_4N5_count = 0
            statsList =[]
            DRTG_cell = 10
            Location_cell = 8
            Rest_cell = 13
            Rest_Detail_cell = 14
            for m in range(row_count-1):
                Total_DRTG = Total_DRTG + tab.cell(row = row_cell, column = DRTG_cell).value
                Total_DRTG_count = Total_DRTG_count + 1
                if tab.cell(row = row_cell, column = Location_cell).value == 'Home':
                    Home_DRTG = Home_DRTG + tab.cell(row = row_cell, column = DRTG_cell).value
                    Home_DRTG_count = Home_DRTG_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 0:
                        Home_DRTG_r0 = Home_DRTG_r0 + tab.cell(row = row_cell, column = DRTG_cell).value
                        Home_DRTG_r0_count = Home_DRTG_r0_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 1:
                        Home_DRTG_r1 = Home_DRTG_r1 + tab.cell(row = row_cell, column = DRTG_cell).value
                        Home_DRTG_r1_count = Home_DRTG_r1_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 2:
                        Home_DRTG_r2 = Home_DRTG_r2 + tab.cell(row = row_cell, column = DRTG_cell).value
                        Home_DRTG_r2_count = Home_DRTG_r2_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value >= 3:
                        Home_DRTG_r3 = Home_DRTG_r3 + tab.cell(row = row_cell, column = DRTG_cell).value
                        Home_DRTG_r3_count = Home_DRTG_r3_count + 1
                    if tab.cell(row = row_cell, column = Rest_Detail_cell).value == '3IN4':
                        Home_DRTG_3N4 = Home_DRTG_3N4 + tab.cell(row = row_cell, column = DRTG_cell).value
                        Home_DRTG_3N4_count = Home_DRTG_3N4_count + 1
                    if tab.cell(row = row_cell, column = Rest_Detail_cell).value == '4IN5':
                        Home_DRTG_4N5 = Home_DRTG_4N5 + tab.cell(row = row_cell, column = DRTG_cell).value
                        Home_DRTG_4N5_count = Home_DRTG_4N5_count + 1
                else:
                    Away_DRTG = Away_DRTG + tab.cell(row = row_cell, column = DRTG_cell).value
                    Away_DRTG_count = Away_DRTG_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 0:
                        Away_DRTG_r0 = Away_DRTG_r0 + tab.cell(row = row_cell, column = DRTG_cell).value
                        Away_DRTG_r0_count = Away_DRTG_r0_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 1:
                        Away_DRTG_r1 = Away_DRTG_r1 + tab.cell(row = row_cell, column = DRTG_cell).value
                        Away_DRTG_r1_count = Away_DRTG_r1_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 2:
                        Away_DRTG_r2 = Away_DRTG_r2 + tab.cell(row = row_cell, column = DRTG_cell).value
                        Away_DRTG_r2_count = Away_DRTG_r2_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value >= 3:
                        Away_DRTG_r3 = Away_DRTG_r3 + tab.cell(row = row_cell, column = DRTG_cell).value
                        Away_DRTG_r3_count = Away_DRTG_r3_count + 1
                    if tab.cell(row = row_cell, column = Rest_Detail_cell).value == '3IN4':
                        Away_DRTG_3N4 = Away_DRTG_3N4 + tab.cell(row = row_cell, column = DRTG_cell).value
                        Away_DRTG_3N4_count = Away_DRTG_3N4_count + 1
                    if tab.cell(row = row_cell, column = Rest_Detail_cell).value == '4IN5':
                        Away_DRTG_4N5 = Away_DRTG_4N5 + tab.cell(row = row_cell, column = DRTG_cell).value
                        Away_DRTG_4N5_count = Away_DRTG_4N5_count + 1
                row_cell = row_cell + 1
            if Total_DRTG_count >0:
                Total_DRTG = Total_DRTG/Total_DRTG_count
            if Home_DRTG_count >0:
                Home_DRTG = Home_DRTG/Home_DRTG_count
            if Home_DRTG_r0_count >0:
                Home_DRTG_r0 = Home_DRTG_r0/Home_DRTG_r0_count
            if Home_DRTG_r1_count >0:
                Home_DRTG_r1 = Home_DRTG_r1/Home_DRTG_r1_count
            if Home_DRTG_r2_count >0:
                Home_DRTG_r2 = Home_DRTG_r2/Home_DRTG_r2_count
            if Home_DRTG_r3_count >0:
                Home_DRTG_r3 = Home_DRTG_r3/Home_DRTG_r3_count
            if Home_DRTG_3N4_count >0:
                Home_DRTG_3N4 = Home_DRTG_3N4/Home_DRTG_3N4_count
            if Home_DRTG_4N5_count >0:
                Home_DRTG_4N5 = Home_DRTG_4N5/Home_DRTG_4N5_count
            if Away_DRTG_count >0:
                Away_DRTG = Away_DRTG/Away_DRTG_count
            if Away_DRTG_r0_count >0:
                Away_DRTG_r0 = Away_DRTG_r0/Away_DRTG_r0_count
            if Away_DRTG_r1_count >0:
                Away_DRTG_r1 = Away_DRTG_r1/Away_DRTG_r1_count
            if Away_DRTG_r2_count >0:
                Away_DRTG_r2 = Away_DRTG_r2/Away_DRTG_r2_count
            if Away_DRTG_r3_count >0:
                Away_DRTG_r3 = Away_DRTG_r3/Away_DRTG_r3_count
            if Away_DRTG_3N4_count >0:
                Away_DRTG_3N4 = Away_DRTG_3N4/Away_DRTG_3N4_count
            if Away_DRTG_4N5_count >0:
                Away_DRTG_4N5 = Away_DRTG_4N5/Away_DRTG_4N5_count
            statsList.append(Total_DRTG)
            statsList.append(Total_DRTG_count)
            statsList.append(Home_DRTG)
            statsList.append(Home_DRTG_count)
            statsList.append(Home_DRTG_r0)
            statsList.append(Home_DRTG_r0_count)
            statsList.append(Home_DRTG_r1)
            statsList.append(Home_DRTG_r1_count)
            statsList.append(Home_DRTG_r2)
            statsList.append(Home_DRTG_r2_count)
            statsList.append(Home_DRTG_r3)
            statsList.append(Home_DRTG_r3_count)
            statsList.append(Home_DRTG_3N4)
            statsList.append(Home_DRTG_3N4_count)
            statsList.append(Home_DRTG_4N5)
            statsList.append(Home_DRTG_4N5_count)
            statsList.append(Away_DRTG)
            statsList.append(Away_DRTG_count)
            statsList.append(Away_DRTG_r0)
            statsList.append(Away_DRTG_r0_count)
            statsList.append(Away_DRTG_r1)
            statsList.append(Away_DRTG_r1_count)
            statsList.append(Away_DRTG_r2)
            statsList.append(Away_DRTG_r2_count)
            statsList.append(Away_DRTG_r3)
            statsList.append(Away_DRTG_r3_count)
            statsList.append(Away_DRTG_3N4)
            statsList.append(Away_DRTG_3N4_count)
            statsList.append(Away_DRTG_4N5)
            statsList.append(Away_DRTG_4N5_count)
            team = j
            tab = wb.get_sheet_by_name('DRTG')
            summary_rows = tab.max_row + 1
            summary_row_cell = 2
            summary_column_cell = 1
            summary_print_column = 2
            print (team)
            for z in range(summary_rows):
                if team == tab.cell(row = summary_row_cell, column = summary_column_cell).value:
                    summary_print_cell = summary_row_cell
                summary_row_cell +=1
            for x in range(len(statsList)):
                tab.cell(row = summary_print_cell, column = summary_print_column).value = statsList[x]
                summary_print_column = summary_print_column + 1
    wb.save('Team_Stats_Raw.xlsx')

def calcPace():
    tabs = wb.get_sheet_names()
    for j in tabs:
        if j != 'Master' and j!= 'ORTG' and j != 'DRTG' and j!= 'Pace':
            tab = wb.get_sheet_by_name(j)
            row_cell = 2
            column_cell = 1
            column_count = tab.max_column
            row_count = tab.max_row
            Total_Pace = 0
            Total_Pace_count = 0
            Home_Pace = 0
            Home_Pace_count = 0
            Home_Pace_r0 = 0
            Home_Pace_r0_count = 0
            Home_Pace_r1 = 0
            Home_Pace_r1_count = 0
            Home_Pace_r2 = 0
            Home_Pace_r2_count = 0
            Home_Pace_r3 = 0
            Home_Pace_r3_count = 0
            Home_Pace_3N4 = 0
            Home_Pace_3N4_count = 0
            Home_Pace_4N5 = 0
            Home_Pace_4N5_count = 0
            Away_Pace = 0
            Away_Pace_count = 0
            Away_Pace_r0 = 0
            Away_Pace_r0_count = 0
            Away_Pace_r1 = 0
            Away_Pace_r1_count = 0
            Away_Pace_r2 = 0
            Away_Pace_r2_count = 0
            Away_Pace_r3 = 0
            Away_Pace_r3_count = 0
            Away_Pace_3N4 = 0
            Away_Pace_3N4_count = 0
            Away_Pace_4N5 = 0
            Away_Pace_4N5_count = 0
            statsList =[]
            Pace_cell = 11
            Location_cell = 8
            Rest_cell = 13
            Rest_Detail_cell = 14
            for m in range(row_count-1):
                Total_Pace = Total_Pace + tab.cell(row = row_cell, column = Pace_cell).value
                Total_Pace_count = Total_Pace_count + 1
                if tab.cell(row = row_cell, column = Location_cell).value == 'Home':
                    Home_Pace = Home_Pace + tab.cell(row = row_cell, column = Pace_cell).value
                    Home_Pace_count = Home_Pace_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 0:
                        Home_Pace_r0 = Home_Pace_r0 + tab.cell(row = row_cell, column = Pace_cell).value
                        Home_Pace_r0_count = Home_Pace_r0_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 1:
                        Home_Pace_r1 = Home_Pace_r1 + tab.cell(row = row_cell, column = Pace_cell).value
                        Home_Pace_r1_count = Home_Pace_r1_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 2:
                        Home_Pace_r2 = Home_Pace_r2 + tab.cell(row = row_cell, column = Pace_cell).value
                        Home_Pace_r2_count = Home_Pace_r2_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value >= 3:
                        Home_Pace_r3 = Home_Pace_r3 + tab.cell(row = row_cell, column = Pace_cell).value
                        Home_Pace_r3_count = Home_Pace_r3_count + 1
                    if tab.cell(row = row_cell, column = Rest_Detail_cell).value == '3IN4':
                        Home_Pace_3N4 = Home_Pace_3N4 + tab.cell(row = row_cell, column = Pace_cell).value
                        Home_Pace_3N4_count = Home_Pace_3N4_count + 1
                    if tab.cell(row = row_cell, column = Rest_Detail_cell).value == '4IN5':
                        Home_Pace_4N5 = Home_Pace_4N5 + tab.cell(row = row_cell, column = Pace_cell).value
                        Home_Pace_4N5_count = Home_Pace_4N5_count + 1
                else:
                    Away_Pace = Away_Pace + tab.cell(row = row_cell, column = Pace_cell).value
                    Away_Pace_count = Away_Pace_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 0:
                        Away_Pace_r0 = Away_Pace_r0 + tab.cell(row = row_cell, column = Pace_cell).value
                        Away_Pace_r0_count = Away_Pace_r0_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 1:
                        Away_Pace_r1 = Away_Pace_r1 + tab.cell(row = row_cell, column = Pace_cell).value
                        Away_Pace_r1_count = Away_Pace_r1_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value == 2:
                        Away_Pace_r2 = Away_Pace_r2 + tab.cell(row = row_cell, column = Pace_cell).value
                        Away_Pace_r2_count = Away_Pace_r2_count + 1
                    if tab.cell(row = row_cell, column = Rest_cell).value >= 3:
                        Away_Pace_r3 = Away_Pace_r3 + tab.cell(row = row_cell, column = Pace_cell).value
                        Away_Pace_r3_count = Away_Pace_r3_count + 1
                    if tab.cell(row = row_cell, column = Rest_Detail_cell).value == '3IN4':
                        Away_Pace_3N4 = Away_Pace_3N4 + tab.cell(row = row_cell, column = Pace_cell).value
                        Away_Pace_3N4_count = Away_Pace_3N4_count + 1
                    if tab.cell(row = row_cell, column = Rest_Detail_cell).value == '4IN5':
                        Away_Pace_4N5 = Away_Pace_4N5 + tab.cell(row = row_cell, column = Pace_cell).value
                        Away_Pace_4N5_count = Away_Pace_4N5_count + 1
                row_cell = row_cell + 1
            if Total_Pace_count >0:
                Total_Pace = Total_Pace/Total_Pace_count
            if Home_Pace_count >0:
                Home_Pace = Home_Pace/Home_Pace_count
            if Home_Pace_r0_count >0:
                Home_Pace_r0 = Home_Pace_r0/Home_Pace_r0_count
            if Home_Pace_r1_count >0:
                Home_Pace_r1 = Home_Pace_r1/Home_Pace_r1_count
            if Home_Pace_r2_count >0:
                Home_Pace_r2 = Home_Pace_r2/Home_Pace_r2_count
            if Home_Pace_r3_count >0:
                Home_Pace_r3 = Home_Pace_r3/Home_Pace_r3_count
            if Home_Pace_3N4_count >0:
                Home_Pace_3N4 = Home_Pace_3N4/Home_Pace_3N4_count
            if Home_Pace_4N5_count >0:
                Home_Pace_4N5 = Home_Pace_4N5/Home_Pace_4N5_count
            if Away_Pace_count >0:
                Away_Pace = Away_Pace/Away_Pace_count
            if Away_Pace_r0_count >0:
                Away_Pace_r0 = Away_Pace_r0/Away_Pace_r0_count
            if Away_Pace_r1_count >0:
                Away_Pace_r1 = Away_Pace_r1/Away_Pace_r1_count
            if Away_Pace_r2_count >0:
                Away_Pace_r2 = Away_Pace_r2/Away_Pace_r2_count
            if Away_Pace_r3_count >0:
                Away_Pace_r3 = Away_Pace_r3/Away_Pace_r3_count
            if Away_Pace_3N4_count >0:
                Away_Pace_3N4 = Away_Pace_3N4/Away_Pace_3N4_count
            if Away_Pace_4N5_count >0:
                Away_Pace_4N5 = Away_Pace_4N5/Away_Pace_4N5_count
            statsList.append(Total_Pace)
            statsList.append(Total_Pace_count)
            statsList.append(Home_Pace)
            statsList.append(Home_Pace_count)
            statsList.append(Home_Pace_r0)
            statsList.append(Home_Pace_r0_count)
            statsList.append(Home_Pace_r1)
            statsList.append(Home_Pace_r1_count)
            statsList.append(Home_Pace_r2)
            statsList.append(Home_Pace_r2_count)
            statsList.append(Home_Pace_r3)
            statsList.append(Home_Pace_r3_count)
            statsList.append(Home_Pace_3N4)
            statsList.append(Home_Pace_3N4_count)
            statsList.append(Home_Pace_4N5)
            statsList.append(Home_Pace_4N5_count)
            statsList.append(Away_Pace)
            statsList.append(Away_Pace_count)
            statsList.append(Away_Pace_r0)
            statsList.append(Away_Pace_r0_count)
            statsList.append(Away_Pace_r1)
            statsList.append(Away_Pace_r1_count)
            statsList.append(Away_Pace_r2)
            statsList.append(Away_Pace_r2_count)
            statsList.append(Away_Pace_r3)
            statsList.append(Away_Pace_r3_count)
            statsList.append(Away_Pace_3N4)
            statsList.append(Away_Pace_3N4_count)
            statsList.append(Away_Pace_4N5)
            statsList.append(Away_Pace_4N5_count)
            team = j
            tab = wb.get_sheet_by_name('Pace')
            summary_rows = tab.max_row + 1
            summary_row_cell = 2
            summary_column_cell = 1
            summary_print_column = 2
            print (team)
            for z in range(summary_rows):
                if team == tab.cell(row = summary_row_cell, column = summary_column_cell).value:
                    summary_print_cell = summary_row_cell
                summary_row_cell +=1
            for x in range(len(statsList)):
                tab.cell(row = summary_print_cell, column = summary_print_column).value = statsList[x]
                summary_print_column = summary_print_column + 1
    wb.save('Team_Stats_Raw.xlsx')

def getSummaryStats():
    calcORTG()
    calcDRTG()
    calcPace()
    wb.save('Team_Stats_Raw.xlsx')
